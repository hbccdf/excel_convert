#! python 3
# -*- coding: utf_8 -*-

import re
from .field import SheetField
from .base import BaseCode
import core.xlproc as xlc
from core.parser import TypeParser
from .code_utils import load_define_data, load_maps
from core.code import GCodeTypeMap
from core.utils import *

sign_re = re.compile(r'^(\w+)<(\w+.*?)>(?:[:](\d+))?$', re.S)


def sign_verify(cell):
    return isinstance(cell, str) and sign_re.match(cell)


class DataStream:
    def __init__(self):
        self._data_bytes = b''

    def write(self, data_bytes):
        self._data_bytes += data_bytes

    @property
    def data_bytes(self):
        return self._data_bytes

    @property
    def length(self):
        return len(self.data_bytes)


class SheetRef:
    def __init__(self, file_path, book_name, sheets):
        self._file_path = file_path
        self._book_name = book_name
        self._sheet_name = sheets[0].name
        self._sheet = sheets[0]
        self._meta_sheets = {}

        self._sub_sheets = []

        for sheet in sheets:
            if self._sheet_name in sheet.name:
                self._sub_sheets.append(sheet)
                continue

            self._meta_sheets[sheet.name] = sheet

    @property
    def file_path(self):
        return self._file_path

    @property
    def book_name(self):
        return self._book_name

    @property
    def name(self):
        return self._sheet_name

    @property
    def sheet(self):
        return self._sheet

    @property
    def meta(self):
        return self._meta_sheets

    @property
    def sub_sheets(self):
        return self._sub_sheets


class SubSheetCode:
    def __init__(self, sheet_code, sheet):
        self._sheet_code = sheet_code
        self._sheet = sheet
        self._name = sheet.name.replace(sheet_code.name, '')
        self._cell2d = None

    @property
    def sub_name(self):
        return self._name

    @property
    def name(self):
        return f'{self._sheet_code.name}{self._name}'

    @property
    def member_name(self):
        return f'{self._sheet_code.member_name}{self._name}'

    @property
    def info(self):
        return f'{self._sheet_code.book_name}>{self.name}'

    @property
    def cell2d(self):
        return self._cell2d

    @property
    def fields(self):
        return self._sheet_code.fields

    @property
    def key_field(self):
        return self._sheet_code.key_field

    @property
    def single(self):
        return self._sheet_code.single

    @property
    def config_type(self):
        if '_Moba' == self.sub_name:
            return 1
        elif '_Diamond' == self.sub_name:
            return 2
        elif '_Fly' == self.sub_name:
            return 3

        return 0

    def load(self):
        if self._sheet_code.ignore:
            return

        n2d, _, _ = xlc.normalize_sheet(self._sheet, sign_verify, self._sheet_code.transpose)
        cells = n2d[1:]

        self._cell2d = cells

        if self._sheet_code.single:
            self._cell2d = cells[:1]


class SheetCode:
    def __init__(self, sheet_ref, test_data):
        # excel
        self._sheet = sheet_ref.sheet
        self._meta_sheets = sheet_ref.meta
        self._file_path = sheet_ref.file_path
        self._book_name = sheet_ref.book_name
        self._sheet_name = sheet_ref.name
        self._sub_sheets = []

        for sub_sheet in sheet_ref.sub_sheets:
            self._sub_sheets.append(SubSheetCode(self, sub_sheet))

        # property
        self._short_name = ''
        self._member_name = ''
        self._key_field = None
        self._fields = []  # SheetField list
        self._self_fields = []
        self._define_fields = []
        self._maps_fields = []

        # base
        self._base_name = ''
        self._base_info = None

        # setting
        self._ignore = False
        self._single = False  # 是否是单行表
        self._transpose = False  # 是否转置
        self._fields_meta = {}

        # data
        self._test_data = None
        self._test_data_new_rows = []
        if test_data and 'test_data' in test_data and 'update' in test_data['test_data']:
            test_data = test_data['test_data']['update']
            if test_data and self.name in test_data:
                self._test_data = test_data[self.name]

    @property
    def name(self):
        return self._sheet_name

    @property
    def short_name(self):
        return self._short_name

    @property
    def member_name(self):
        return self._member_name

    @property
    def book_name(self):
        return self._book_name

    @property
    def info(self):
        return f'{self._book_name}>{self.name}'

    @property
    def ignore(self):
        return self._ignore

    @property
    def single(self):
        return self._single

    @property
    def transpose(self):
        return self._transpose

    @property
    def sub_sheets(self):
        return self._sub_sheets

    @property
    def key_field(self):
        return self._key_field

    @property
    def fields(self):
        return self._fields

    @property
    def self_fields(self):
        return self._self_fields

    @property
    def base_name(self):
        return self._base_name

    def get_sub_sheet_by_config_type(self, config_type):
        for sub_sheet in self.sub_sheets:
            if sub_sheet.config_type == config_type:
                return sub_sheet

        return self.sub_sheets[0]

    def load_define(self, define_data):
        self._load_define_data(define_data)

        define_sheet = self._get_sheet('define')
        if not define_sheet:
            return

        a2d = xlc.array2d(define_sheet)
        TypeParser.parse(a2d)

    def load(self, bases):
        # 二次处理
        self._load()

        if self._ignore:
            return

        for f in self._maps_fields:
            f.parse_maps()

        self._fields.extend(self._maps_fields)
        self._fields.sort(key=lambda elem: elem.position)

        if 'Config' not in self._sheet_name:
            self._short_name = self.name + 'Cfg'
        else:
            self._short_name = self.name.replace('Config', 'Cfg')

        if self._single:
            self._member_name = f'{self.name}Info'
        else:
            self._member_name = f'{self.name}Map'

        for f in self._fields:
            if f.is_key:
                self._key_field = f
                break

        if self._base_name != '':
            if self._base_name not in bases:
                bases[self._base_name] = BaseCode(self._base_name)

            self._base_info = bases[self._base_name]
            self._base_info.add_fields(self._sheet_name, self._fields)

    def process(self):
        self._process_base_info()
        self._process_define_fields()

    def _get_sheet(self, name):
        if name in self._meta_sheets:
            return self._meta_sheets[name]
        return None

    def _get_test_data(self, field):
        if not field:
            return None, 0

        if not self.single and field == self._key_field:
            return None, 0

        if not self._test_data or len(self._test_data) == 0:
            return None, 0

        if field.name not in self._test_data:
            return None, 0

        field_test_data = self._test_data[field.name]
        field_current_key = self._key_field.value_data
        # 以字段为主，然后再寻找id的方式
        if field_test_data is not None:
            if self.single:
                return field_test_data, 0

            if field_current_key is None:
                field_current_key = 0

            if field_current_key and field_current_key in field_test_data:
                return field_test_data[field_current_key], field_current_key

            if 'all' in field_test_data:
                return field_test_data['all'], field_current_key

        # 以id为主，然后再寻找字段的方式
        elif field_current_key is not None and field_current_key in self._test_data:
            row_test_data = self._test_data[field_current_key]
            if row_test_data is None:
                row_test_data = self._test_data['all']

            if row_test_data and field.name in row_test_data:
                return row_test_data[field.name], field_current_key

        return None, 0

    def _load_define_data(self, define_data):
        fields = load_define_data(self, define_data)
        self._define_fields.extend(fields)

        maps_fields = load_maps(self, define_data)
        self._maps_fields.extend(maps_fields)

    def _load(self):
        self._load_setting()

        if self._ignore:
            return

        n2d, a2d, sign_pos = xlc.normalize_sheet(self._sheet, sign_verify, self._transpose)

        desc_list = a2d[sign_pos.r - 1]
        title = n2d[0]

        for i, sign in enumerate(title):
            (r, c, s) = sign

            m = sign_re.match(s)
            name = m[1]
            sign = m[2]
            position = (i + 1) * 10
            if m[3]:
                position = m[3]

            desc = desc_list[i]
            meta_position, setting = get_key_value(self._fields_meta, name, (0, ''))
            if meta_position > 0:
                position = meta_position

            f = SheetField(self, name, sign, desc, i, position, setting)

            if len(self._fields) == 0:
                f.is_key = True

            if f.ignore:
                print(f'ignore column : {f.info}')
                continue

            self._fields.append(f)

        for sub_sheet in self._sub_sheets:
            sub_sheet.load()
            self._load_test_data(sub_sheet.cell2d)

    def _load_setting(self):
        self._load_sheet_setting()
        self._load_meta_setting()

    def _load_sheet_setting(self):
        setting_sheet = self._get_sheet('setting')
        if not setting_sheet:
            return

        setting_a2d = xlc.array2d(setting_sheet)
        setting_rows = setting_a2d[1:]
        for row in setting_rows:
            key = get_list_value(row, 0, None)
            val = get_list_value(row, 1, None)
            languages = get_list_value(row, 2, None)

            if not key:
                continue

            if not is_valid_language(languages):
                continue

            self._process_setting(key, val)

    def _load_meta_setting(self):
        meet_sheet = self._get_sheet('meta')
        if not meet_sheet:
            return

        meta_a2d = xlc.array2d(meet_sheet)
        meta_rows = meta_a2d[1:]
        for row in meta_rows:
            if len(row) == 0:
                continue

            key = get_list_value(row, 0, None)
            position = get_list_value(row, 1, 0)
            languages = get_list_value(row, 2, None)
            setting_str = get_list_value(row, 3, '')
            if not is_valid_language(languages):
                continue

            if len(row) >= 4:
                setting_str = '|'.join(row[3:])

            self._fields_meta[key] = (position, setting_str)

    def _process_setting(self, key, val):
        if key == 'single':
            self._single = True
        elif key == 'transpose':
            self._transpose = True
        elif key == 'language':
            self._ignore = not is_valid_language(val)
        elif key == 'base':
            self._base_name = val

    def _load_test_data(self, cell2d):
        if not self._test_data or 'new' not in self._test_data:
            return

        if len(cell2d) <= 0:
            return

        last_row = cell2d[-1]
        last_row_index, _, _ = last_row[0]

        new_row_count = 0
        row_index = last_row_index + 1
        row_strings = self._test_data['new']
        for row_string in row_strings:
            row = None
            if isinstance(row_string, list):
                row = row_string
            elif isinstance(row_string, str):
                row = parse_value_list(row_string)

            if not row or len(row) == 0:
                continue

            row = [(row_index, i, value) for i, value in enumerate(row)]
            cell2d.append(row)
            self._test_data_new_rows.append(row)
            new_row_count += 1
            row_index += 1

        if new_row_count > 0:
            print(f'new row {new_row_count}')

    def write_data(self, writer, show_test_data=True):
        for sub_sheet in self._sub_sheets:
            cell2d = sub_sheet.cell2d

            row_len = len(cell2d)
            writer.write_sheet_begin(sub_sheet, row_len)

            key_ids = []

            for i, row in enumerate(cell2d):
                for field in self._fields:
                    if field.only_define or field.is_value_map:
                        continue

                    column_index = field.column_index
                    (r, c, data) = row[column_index]

                    field_test_data, field_key = self._get_test_data(field)
                    if field_test_data is not None:
                        data = field_test_data
                        if show_test_data:
                            print(f'test_data: {sub_sheet.info}>{field.name},{field_key}: {field_test_data}')

                    ok, info = field.set_data(data)
                    if not ok:
                        raise Exception(f'sheet info: {sub_sheet.info}>{field.name},{i}: {info}')

                for field in self._maps_fields:
                    ok, info = field.update_value_map_data()
                    if not ok:
                        raise Exception(f'sheet info: {sub_sheet.info}>{field.name},{i}: {info}')

                key_id = sub_sheet.key_field.value_data
                if key_id in key_ids:
                    raise Exception(f'sheet info: {sub_sheet.info}>row {i}: repeat key id {key_id}')
                else:
                    key_ids.append(key_id)

                writer.write_row_begin(sub_sheet, i, row_len)

                for field in self._fields:
                    if field.only_define or field.in_value_map:
                        continue
                    writer.write_field_begin(field)
                    writer.write_field_end(field)

                writer.write_row_end(sub_sheet, i, row_len)

            writer.write_sheet_end(sub_sheet)

    def update_data(self):
        if not self._test_data or len(self._test_data) == 0:
            return

        import openpyxl as xl
        wb = xl.load_workbook(filename=self._file_path)
        if self.name not in wb:
            return

        is_modify = False

        for sub_sheet in self._sub_sheets:
            cell2d = sub_sheet.cell2d

            wb_sheet = wb[sub_sheet.name]

            for i, row in enumerate(cell2d):
                for field in self._fields:
                    if field.only_define or field.is_value_map:
                        continue

                    column_index = field.column_index
                    (r, c, data) = row[column_index]
                    if field == self._key_field:
                        field.set_data(data)
                        continue

                    field_test_data, field_key = self._get_test_data(field)
                    if field_test_data is None:
                        continue

                    data = field_test_data
                    print(f'update test_data: {sub_sheet.info}>{field.name},{field_key}: {field_test_data}')
                    wb_sheet.cell(r + 1, c + 1, data)
                    is_modify = True

            for row in self._test_data_new_rows:
                print(f'new row: {sub_sheet.info}: {row}')
                for row_value in row:
                    r, c, data = row_value
                    wb_sheet.cell(r + 1, c + 1, data)
                    is_modify = True

        if is_modify:
            wb.save(self._file_path)

    def _process_base_info(self):
        if not self._base_info:
            self._self_fields = self._fields
            return

        for f in self._fields:
            if f not in self._base_info:
                self._self_fields.append(f)

    def _process_define_fields(self):
        self._self_fields.extend(self._define_fields)


class AllCode:
    def __init__(self, sheet_refs, define_file, test_data_file, namespace):
        self._sheets = []
        self._bases = {}
        self._singles = []
        self._not_singles = []
        self._namespace = namespace

        print(define_file)
        self._defines = load_yaml_data(define_file)

        self._test_data = None
        if test_data_file and os.path.exists(test_data_file):
            self._test_data = load_yaml_data(test_data_file)

        TypeParser.parse_define_data(self._defines)

        sheets = [SheetCode(sheet, self._test_data) for sheet in sheet_refs]

        for sheet in sheets:
            print(f'load define: {sheet.info}')
            sheet.load_define(self._defines)

        for sheet in sheets:
            print(f'load excel data: {sheet.info}')
            sheet.load(self._bases)

        for _, base in self._bases.items():
            base.process(self._defines)

        for sheet in sheets:
            sheet.process()

        for sheet in sheets:
            if sheet.ignore:
                continue

            self._sheets.append(sheet)
            if not sheet.single:
                self._not_singles.append(sheet)
            else:
                self._singles.append(sheet)

    @property
    def sheets(self):
        return self._sheets

    @property
    def bases(self):
        return self._bases

    @property
    def enums(self):
        return GCodeTypeMap.enums

    @property
    def structs(self):
        return GCodeTypeMap.structs

    @property
    def singles(self):
        return self._singles

    @property
    def not_singles(self):
        return self._not_singles

    @property
    def sheets_count(self):
        count = 0
        for sheet in self._sheets:
            count += len(sheet.sub_sheets)

        return count

    @property
    def namespace(self):
        return self._namespace